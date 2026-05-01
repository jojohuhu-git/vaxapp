#!/usr/bin/env python3
"""Dump all in-scope CDSI 4.6 Excel sheets to a single flat JSON file.

Deterministic — no interpretation, just raw cell data. The output feeds the
Step 2b interpreter (LLM) which turns the flat dump into structured rules.

Usage:
    python3 scripts/dump-cdsi-excel.py
"""
import json
import os
from pathlib import Path

import openpyxl

EXCEL_DIR = Path("/Users/joannehuang/Downloads/Version 4.64 - 508/Excel")
OUT_FILE = Path(__file__).resolve().parent.parent / "src" / "data" / "cdsi-4.6-raw.json"

# In-scope antigens → source files
ANTIGEN_FILES = {
    "DTaP":    ["AntigenSupportingData- Pertussis-508.xlsx",
                "AntigenSupportingData- Diphtheria-508.xlsx",
                "AntigenSupportingData- Tetanus-508.xlsx"],
    "Tdap":    ["AntigenSupportingData- Pertussis-508.xlsx",
                "AntigenSupportingData- Diphtheria-508.xlsx",
                "AntigenSupportingData- Tetanus-508.xlsx"],
    "Td":      ["AntigenSupportingData- Diphtheria-508.xlsx",
                "AntigenSupportingData- Tetanus-508.xlsx"],
    "MenACWY": ["AntigenSupportingData- Meningococcal-508.xlsx"],
    "MenB":    ["AntigenSupportingData- Meningococcal B-508.xlsx"],
    "HepB":    ["AntigenSupportingData- HepB-508.xlsx"],
    "IPV":     ["AntigenSupportingData- Polio-508.xlsx"],
    "Hib":     ["AntigenSupportingData- Hib-508.xlsx"],
    "PCV":     ["AntigenSupportingData- Pneumococcal-508.xlsx"],
    "PPSV23":  ["AntigenSupportingData- Pneumococcal-508.xlsx"],
    # Round-2 audit additions (2026-04-30):
    "RV":      ["AntigenSupportingData- Rotavirus-508.xlsx"],
    "MMR":     ["AntigenSupportingData- Measles-508.xlsx",
                "AntigenSupportingData- Mumps-508.xlsx",
                "AntigenSupportingData- Rubella-508.xlsx"],
    "VAR":     ["AntigenSupportingData- Varicella-508.xlsx"],
    "HepA":    ["AntigenSupportingData- HepA-508.xlsx"],
    "HPV":     ["AntigenSupportingData- HPV-508.xlsx"],
    "RSV":     ["AntigenSupportingData- RSV-508.xlsx"],
    "Flu":     ["AntigenSupportingData- Influenza-508.xlsx"],
    "COVID":   ["AntigenSupportingData- COVID-19-508.xlsx"],
}

# Cross-reference files (loaded once, shared)
SCHEDULE_FILES = [
    "ScheduleSupportingData- CVX to Antigen Map-508.xlsx",
    "ScheduleSupportingData- Vaccine Group-508.xlsx",
    "ScheduleSupportingData- Vaccine Group to Antigen Map-508.xlsx",
    "ScheduleSupportingData- Live Virus Conflicts-508.xlsx",
]

# Sheets to skip everywhere — pure metadata
SKIP_SHEETS = {"Change History", "FAQ"}


def dump_sheet(ws):
    """Return a list-of-rows representation of a worksheet."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Trim trailing Nones for compactness; keep leading/internal Nones
        last = len(row)
        while last > 0 and row[last - 1] is None:
            last -= 1
        if last == 0:
            continue
        rows.append(list(row[:last]))
    return rows


def dump_workbook(path):
    """Return {sheetName: [rows]} for one workbook."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        out[sheet_name] = dump_sheet(ws)
    wb.close()
    return out


def main():
    # Cache: load each unique workbook only once (some are referenced by multiple antigens)
    workbook_cache = {}

    def load(file_name):
        if file_name not in workbook_cache:
            full = EXCEL_DIR / file_name
            if not full.exists():
                print(f"WARN: missing file {full}")
                workbook_cache[file_name] = None
            else:
                print(f"  loading {file_name} …")
                workbook_cache[file_name] = dump_workbook(full)
        return workbook_cache[file_name]

    out = {
        "version": "4.6",
        "source": str(EXCEL_DIR),
        "convention_notes": [
            "All cell values are the raw values from the .xlsx (data_only=True so formulas are evaluated).",
            "Empty trailing cells in each row are trimmed; internal blanks are preserved as None/null.",
            "Sheets named 'Change History' and 'FAQ' are skipped.",
            "Some antigen files (Pertussis, Diphtheria, Tetanus) are shared by DTaP/Tdap/Td — each antigen entry below references all relevant source files."
        ],
        "antigens": {},
        "scheduleSupportingData": {},
    }

    print("== Antigen files ==")
    for antigen, files in ANTIGEN_FILES.items():
        print(f"{antigen}:")
        out["antigens"][antigen] = {
            "sourceFiles": files,
            "workbooks": {f: load(f) for f in files},
        }

    print("== Schedule supporting data ==")
    for f in SCHEDULE_FILES:
        out["scheduleSupportingData"][f] = load(f)

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_FILE, "w") as fh:
        json.dump(out, fh, indent=1, default=str)
    size_kb = OUT_FILE.stat().st_size / 1024
    print(f"\nWrote {OUT_FILE} ({size_kb:.0f} KB)")


if __name__ == "__main__":
    main()
